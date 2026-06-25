"""
Export service for Excel, PDF, and Markdown file generation.
"""

import io
from datetime import datetime

from app.models.schemas import GenerationResult, RequirementType
from app.utils.logger import get_logger

logger = get_logger("export")


class ExportService:
    """Service for exporting generated requirements to various formats."""

    def export_to_excel(self, result: GenerationResult) -> bytes:
        """Export all requirements to a styled Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        critical_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        high_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        medium_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        def write_header(ws, headers):
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = thin_border

        def priority_fill(p):
            if p in ("critical", "blocker"):
                return critical_fill
            elif p == "high":
                return high_fill
            return medium_fill

        def auto_width(ws):
            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

        # ─── Functional Requirements ───
        ws = wb.active
        ws.title = "Functional Requirements"
        write_header(ws, ["ID", "Module", "Description", "Acceptance Criteria", "Source Files", "Priority", "Severity"])
        for i, r in enumerate(result.functional_requirements, 2):
            ws.cell(row=i, column=1, value=r.requirement_id).border = thin_border
            ws.cell(row=i, column=2, value=r.module).border = thin_border
            ws.cell(row=i, column=3, value=r.description).border = thin_border
            ws.cell(row=i, column=4, value=r.acceptance_criteria).border = thin_border
            ws.cell(row=i, column=5, value="; ".join(r.source_files)).border = thin_border
            c = ws.cell(row=i, column=6, value=r.priority.value)
            c.border = thin_border
            c.fill = priority_fill(r.priority.value)
            ws.cell(row=i, column=7, value=r.severity.value).border = thin_border
        auto_width(ws)

        # ─── API Requirements ───
        if result.api_requirements:
            ws2 = wb.create_sheet("API Requirements")
            write_header(ws2, ["ID", "Module", "Endpoint Description", "Acceptance Criteria", "Source Files", "Priority"])
            for i, r in enumerate(result.api_requirements, 2):
                ws2.cell(row=i, column=1, value=r.requirement_id).border = thin_border
                ws2.cell(row=i, column=2, value=r.module).border = thin_border
                ws2.cell(row=i, column=3, value=r.description).border = thin_border
                ws2.cell(row=i, column=4, value=r.acceptance_criteria).border = thin_border
                ws2.cell(row=i, column=5, value="; ".join(r.source_files)).border = thin_border
                c = ws2.cell(row=i, column=6, value=r.priority.value)
                c.border = thin_border
                c.fill = priority_fill(r.priority.value)
            auto_width(ws2)

        # ─── User Stories ───
        if result.user_stories:
            ws3 = wb.create_sheet("User Stories")
            write_header(ws3, ["ID", "Module", "Persona", "Action", "Benefit", "Acceptance Criteria", "Priority"])
            for i, s in enumerate(result.user_stories, 2):
                ws3.cell(row=i, column=1, value=s.story_id).border = thin_border
                ws3.cell(row=i, column=2, value=s.module).border = thin_border
                ws3.cell(row=i, column=3, value=s.persona).border = thin_border
                ws3.cell(row=i, column=4, value=s.action).border = thin_border
                ws3.cell(row=i, column=5, value=s.benefit).border = thin_border
                ws3.cell(row=i, column=6, value="; ".join(s.acceptance_criteria)).border = thin_border
                ws3.cell(row=i, column=7, value=s.priority.value).border = thin_border
            auto_width(ws3)

        # ─── Validation Rules ───
        if result.validation_rules:
            ws4 = wb.create_sheet("Validation Rules")
            write_header(ws4, ["ID", "Module", "Field", "Rule", "Constraint Type", "Priority"])
            for i, v in enumerate(result.validation_rules, 2):
                ws4.cell(row=i, column=1, value=v.rule_id).border = thin_border
                ws4.cell(row=i, column=2, value=v.module).border = thin_border
                ws4.cell(row=i, column=3, value=v.field_or_parameter).border = thin_border
                ws4.cell(row=i, column=4, value=v.rule_description).border = thin_border
                ws4.cell(row=i, column=5, value=v.constraint_type).border = thin_border
                ws4.cell(row=i, column=6, value=v.priority.value).border = thin_border
            auto_width(ws4)

        # ─── Test Cases ───
        if result.test_cases:
            ws5 = wb.create_sheet("Test Cases")
            write_header(ws5, ["ID", "Module", "Scenario", "Description", "Input", "Expected Output", "Edge Case", "Related Req"])
            for i, t in enumerate(result.test_cases, 2):
                ws5.cell(row=i, column=1, value=t.test_id).border = thin_border
                ws5.cell(row=i, column=2, value=t.module).border = thin_border
                ws5.cell(row=i, column=3, value=t.scenario).border = thin_border
                ws5.cell(row=i, column=4, value=t.description).border = thin_border
                ws5.cell(row=i, column=5, value=t.test_input).border = thin_border
                ws5.cell(row=i, column=6, value=t.expected_output).border = thin_border
                ws5.cell(row=i, column=7, value="Yes" if t.edge_case else "No").border = thin_border
                ws5.cell(row=i, column=8, value=t.related_requirement).border = thin_border
            auto_width(ws5)

        # ─── Edge Cases ───
        if result.edge_cases:
            ws6 = wb.create_sheet("Edge Cases")
            write_header(ws6, ["ID", "Module", "Scenario", "Description", "Boundary", "Expected Behavior", "Severity"])
            for i, e in enumerate(result.edge_cases, 2):
                ws6.cell(row=i, column=1, value=e.edge_case_id).border = thin_border
                ws6.cell(row=i, column=2, value=e.module).border = thin_border
                ws6.cell(row=i, column=3, value=e.scenario).border = thin_border
                ws6.cell(row=i, column=4, value=e.description).border = thin_border
                ws6.cell(row=i, column=5, value=e.boundary_condition).border = thin_border
                ws6.cell(row=i, column=6, value=e.expected_behavior).border = thin_border
                ws6.cell(row=i, column=7, value=e.severity.value).border = thin_border
            auto_width(ws6)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_to_markdown(self, result: GenerationResult) -> str:
        """Export all requirements to GitHub-flavored Markdown."""
        lines = [f"# Requirements Document — {result.repo_summary.repo_name if result.repo_summary else 'Repository'}",
                 f"\n_Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}_\n"]

        if result.functional_requirements:
            lines.append("## Functional Requirements\n")
            lines.append("| ID | Module | Description | Priority | Severity |")
            lines.append("|---|---|---|---|---|")
            for r in result.functional_requirements:
                lines.append(f"| {r.requirement_id} | {r.module} | {r.description} | {r.priority.value} | {r.severity.value} |")
            lines.append("")

        if result.api_requirements:
            lines.append("## API Requirements\n")
            lines.append("| ID | Module | Endpoint | Priority |")
            lines.append("|---|---|---|---|")
            for r in result.api_requirements:
                lines.append(f"| {r.requirement_id} | {r.module} | {r.description} | {r.priority.value} |")
            lines.append("")

        if result.user_stories:
            lines.append("## User Stories\n")
            for s in result.user_stories:
                lines.append(f"### {s.story_id}: {s.module}")
                lines.append(f"**As a** {s.persona}, **I want** {s.action}, **so that** {s.benefit}\n")
                if s.acceptance_criteria:
                    lines.append("**Acceptance Criteria:**")
                    for ac in s.acceptance_criteria:
                        lines.append(f"- {ac}")
                lines.append("")

        if result.validation_rules:
            lines.append("## Validation Rules\n")
            lines.append("| ID | Module | Field | Rule | Type |")
            lines.append("|---|---|---|---|---|")
            for v in result.validation_rules:
                lines.append(f"| {v.rule_id} | {v.module} | {v.field_or_parameter} | {v.rule_description} | {v.constraint_type} |")
            lines.append("")

        if result.test_cases:
            lines.append("## Unit Test Cases\n")
            lines.append("| ID | Module | Scenario | Input | Expected Output | Edge Case |")
            lines.append("|---|---|---|---|---|---|")
            for t in result.test_cases:
                ec = "✓" if t.edge_case else ""
                lines.append(f"| {t.test_id} | {t.module} | {t.scenario} | {t.test_input} | {t.expected_output} | {ec} |")
            lines.append("")

        if result.edge_cases:
            lines.append("## Edge Cases\n")
            lines.append("| ID | Module | Scenario | Boundary | Expected Behavior | Severity |")
            lines.append("|---|---|---|---|---|---|")
            for e in result.edge_cases:
                lines.append(f"| {e.edge_case_id} | {e.module} | {e.scenario} | {e.boundary_condition} | {e.expected_behavior} | {e.severity.value} |")

        return "\n".join(lines)

    def export_to_pdf(self, result: GenerationResult) -> bytes:
        """Export requirements to a PDF report."""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.units import inch

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=0.5*inch, rightMargin=0.5*inch)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=18, textColor=colors.HexColor("#4338CA"))
        heading_style = ParagraphStyle("Heading", parent=styles["Heading2"], fontSize=14, textColor=colors.HexColor("#1E293B"))
        cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=8, leading=10)

        elements = []
        repo_name = result.repo_summary.repo_name if result.repo_summary else "Repository"
        elements.append(Paragraph(f"Requirements Document — {repo_name}", title_style))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
        elements.append(Spacer(1, 24))

        table_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4338CA")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ])

        def p(text):
            return Paragraph(str(text)[:200], cell_style)

        if result.functional_requirements:
            elements.append(Paragraph("Functional Requirements", heading_style))
            elements.append(Spacer(1, 8))
            data = [["ID", "Module", "Description", "Priority", "Severity"]]
            for r in result.functional_requirements:
                data.append([r.requirement_id, r.module, p(r.description), r.priority.value, r.severity.value])
            t = Table(data, colWidths=[60, 80, 400, 60, 60])
            t.setStyle(table_style)
            elements.append(t)
            elements.append(Spacer(1, 20))

        if result.test_cases:
            elements.append(PageBreak())
            elements.append(Paragraph("Unit Test Cases", heading_style))
            elements.append(Spacer(1, 8))
            data = [["ID", "Module", "Scenario", "Expected Output", "Edge"]]
            for tc in result.test_cases:
                data.append([tc.test_id, tc.module, p(tc.scenario), p(tc.expected_output), "Y" if tc.edge_case else ""])
            t = Table(data, colWidths=[50, 70, 250, 250, 30])
            t.setStyle(table_style)
            elements.append(t)

        doc.build(elements)
        return output.getvalue()


def get_export_service() -> ExportService:
    return ExportService()
